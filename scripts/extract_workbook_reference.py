from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "data/model/accessibility_business_case.xlsx"
OUTPUT = ROOT / "data/reference"


def main() -> None:
    workbook = load_workbook(WORKBOOK, data_only=True, read_only=True)
    OUTPUT.mkdir(parents=True, exist_ok=True)

    case = workbook["Case_Data"]
    case_rows = list(case.iter_rows(min_row=2, max_row=5, values_only=True))
    pd.DataFrame(
        [{"page": row[0], "annual_views": row[13], "estimated_issues": row[15]} for row in case_rows]
    ).to_csv(OUTPUT / "case_pages.csv", index=False)

    mcda = workbook["MCDA_Prioritisation"]
    mcda_rows = list(mcda.iter_rows(min_row=2, max_row=5, values_only=True))
    pd.DataFrame(
        [
            {
                "page": row[0], "views": row[1], "issues": row[2],
                "traffic_score": row[3], "issue_score": row[4],
                "process_criticality": row[5], "legal_hr_sensitivity": row[6],
                "feasibility": row[7], "reported_score": row[8], "reported_rank": row[9],
            }
            for row in mcda_rows
        ]
    ).to_csv(OUTPUT / "mcda_priorities.csv", index=False)

    backlog = workbook["Audit_Backlog"]
    headers = [cell.value for cell in next(backlog.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in backlog.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        rows.append(dict(zip(headers[:12], row[:12], strict=True)))
    pd.DataFrame(rows).to_csv(OUTPUT / "audit_backlog.csv", index=False)

    monte = workbook["Monte_Carlo"]
    simulations = []
    for row in monte.iter_rows(min_row=3, max_row=1002, values_only=True):
        if row[0] is None:
            continue
        simulations.append({"iteration": row[0], "npv": row[14], "positive_npv": row[15]})
    pd.DataFrame(simulations).to_csv(OUTPUT / "monte_carlo_npv.csv", index=False)
    print(f"Extracted {len(simulations)} frozen simulations")


if __name__ == "__main__":
    main()
